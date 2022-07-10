# 預先整理訓練數據集question和answers的list
import json
import time
#from numba import jit
import openpyxl


'''操作xlsx格式的表格文件'''
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)

workbook = None
q_list = []
a_list = []


'''###1
#如果要訓練模型，更換.bin,C:\\Users\\User\\Desktop\\Bert訓練包\\Bert_project\\bert\\data\\DRCD_\\DRCD_training.json', 'r',encoding="utf-8的資料集也要更換
with open('.\\data\\DRCD_\\DRCD_training.json', 'r',encoding="utf-8") as file:
    
    st_ = file.read()
    row_data = json.loads(st_)
    
    # 442个主题,逐一个主题检视
    title_count = 0
    while (title_count < len(row_data['data'])):
        title_data = row_data['data'][title_count]
        
        # 66个段落,逐一个段落检视
        paragraphs_count = 0
        while (paragraphs_count < len(title_data['paragraphs'])):
            paragraphs_data = title_data['paragraphs'][paragraphs_count]
            
            # 15个问题,逐一个问题检视
            qas_count = 0
            while (qas_count < len(paragraphs_data['qas'])):
                qas_data = paragraphs_data['qas'][qas_count]
                
                # 取出question, answers组成数列
                
                
                # 假若没有question则自己补上no question
                if qas_data['question'] == []:
                    question_data_clean = qas_data['question'] + [' no question']
                else:
                    question_data_clean = qas_data['question']
                q_list = q_list + [question_data_clean]
                

                # 假若没有answers则自己补上no answers    
                if qas_data['answers'] == []:
                    answers_data_clean = qas_data['answers'] + [' no answers']
                else:
                    answers_data_clean = qas_data['answers'][0]['text']
                a_list = a_list + [answers_data_clean]
                
                
                qas_count = qas_count + 1
                
            paragraphs_count = paragraphs_count + 1
            
        title_count = title_count + 1
    

# question和answers存成json
file_name_1 = 'q_list.json'
file_name_2 = 'a_list.json'
with open(file_name_1,'w') as file_object:
    json.dump(q_list,file_object)
with open(file_name_2,'w') as file_object:
    json.dump(a_list,file_object)
'''###1
    
#########    
#讀取完成後檢查是否每一個問題都有對應到答案
def read_corpus():
    """
    读取给定的语料库，并把问题列表和答案列表分别写入到 qlist, alist 里面。 在此过程中，不用对字符换做任何的处理（这部分需要在 Part 2.3里处理）
    qlist = ["问题1"， “问题2”， “问题3” ....]
    alist = ["答案1", "答案2", "答案3" ....]
    务必要让每一个问题和答案对应起来（下标位置一致）
    """
    # TODO 需要完成的代码部分 ...
    
    # 读取question的json
    with open('q_list.json', 'r') as file:
        question = file.read()
        qlist = json.loads(question)

    # 读取answers的json
    with open('a_list.json', 'r') as file:
        answers = file.read()
        alist = json.loads(answers)
    
    assert len(qlist) == len(alist)  # 确保长度一样
    return qlist, alist

corpus=read_corpus()
qlist=corpus[0]
alist=corpus[1]
#print(qlist)


'''####2
##########
# 计算单词
#import nltk #英文
import jieba
word_list = []   # 计算不重复的单词
word_total_list = []   # 计算总共的单词
for i in qlist:
    #print(i)
    #words = nltk.word_tokenize(i)
    words=list(jieba.cut_for_search(i))
    #print(words)
    
    # 计算不重复的单词 
    word_list = word_list + words
    #print(word_list)
    word_list = list(set(word_list))
    #print(word_list)
    
    # 计算总共的单词
    word_total_list = word_total_list + words
'''####2

#-----------------------------------------------------------------------------------------------------------------


'''####3
# question的word_list和word_total_list存成json
file_name_3 = 'q_word_list.json'
file_name_4 = 'q_word_total_list.json'
with open(file_name_3,'w') as file_object:
    json.dump(word_list,file_object)
with open(file_name_4,'w') as file_object:
    json.dump(word_total_list,file_object)
'''####3



'''####4
import collections

# 计算每个词的出现个数,变成字典形式,依照出现个数多的降幂排列
frequency_count = collections.Counter(word_total_list)

# 计算单词出现的次数
## 先取出字典每个单词出现的次数
dict_value = frequency_count.values()
#print(dict_value)

## 再计算出现的单词次数的个数
word_count_dict_temp = collections.Counter(dict_value)
### 变成items后依照key值排序
word_count_items = sorted(word_count_dict_temp.items(),key = lambda item:item[0])

word_count =[]
word_number_count = []

for i,j in word_count_items:
    word_count += [i]
    word_number_count += [j]

    
#print(word_count)
#print(word_number_count)


#-----------------------------------------------------------------------------------------------------------------

import matplotlib.pyplot as plt

# 画出长条图,因为数量太多无法显示,故仅显示前20数据
#plt.axis([0, 1.5,0, 600])
plt.bar(word_count[:30], word_number_count[:30])
plt.show()
'''####4

#########
#@jit()
def Read_stopwords():
    fb=open(".\stopwords\stopwords.txt","r",encoding='utf-8')
    stopwords_list=[]
    for stopwords in fb.readlines():
        stopwords=stopwords.strip("\n")
        stopwords_list.append(stopwords)
    return stopwords_list
stopwords_list=Read_stopwords()


#去除停用詞
#逐一将每个问题拆解成一个一个单字,并去除停用词,标点符号
def movestopwords(sentence):
    #from nltk.corpus import stopwords
    #from nltk.tokenize import word_tokenize
    
    q_filtered_list = []
    # 将每一个问题取出来
    for question in sentence:
        
        # 把问题拆解成一个一个单词
        word_tokens = list(jieba.cut_for_search(question))
        
        
        # 过滤掉停用词
        ## 增加stopwords
        #stop_words = set(stopwords.words('chinese'))
        #stop_words.update(['.', ',', '"', "'", '?', '!', ':', ';', '(', ')', '[', ']', '{', '}'])
        #filtered_words = [word for word in word_tokens if word not in stop_words]
        filtered_words =[word for word in word_tokens if word not in stopwords_list]
        #print(filtered_words)

        # 将单词组合成句子
        filtered_sentens = ' '.join(filtered_words)
       
        q_filtered_list += [filtered_sentens]

    return q_filtered_list
#movestopwords(["廣州的快速公交運輸系統每多久就會有一輛巴士？"])
#print(movestopwords(["廣州的快速公交運輸系統每多久就會有一輛巴士？"]))


'''#####5
import json

with open('q_list.json', 'r') as file:
    question = file.read()
    q_list = json.loads(question)

# 英文的大写全部改成小写
q_list_lower_case = [item.lower() for item in q_list]

# 整理完的问题集
clean_q_list = movestopwords(q_list_lower_case)
#print(clean_q_list)
#----------------------------------------------------------------------------------------------------------------- 

# question的clean_q_list存成json
file_name_5 = 'clean_q_word_list.json'
with open(file_name_5,'w') as file_object:
    json.dump(clean_q_list,file_object)
'''#####5


###########
import torch
from pytorch_pretrained_bert import BertTokenizer, BertModel
#from pytorch_pretrained_bert import BertForMaskedLM

# 有5种预训练好的bert模型可以参考
## bert-base-uncased: 12-layer, 768-hidden, 12-heads, 110M parameters
## bert-large-uncased: 24-layer, 1024-hidden, 16-heads, 340M parameters
## bert-base-cased: 12-layer, 768-hidden, 12-heads , 110M parameters
## bert-base-multilingual: 102 languages, 12-layer, 768-hidden, 12-heads, 110M parameters
## bert-base-chinese: Chinese Simplified and Traditional, 12-layer, 768-hidden, 12-heads, 110M parameters

# 使用bert-base-uncased
tokenizer = BertTokenizer.from_pretrained('.\model') #.\model
model = BertModel.from_pretrained('.\model') #.\model
#tokenizer = BertTokenizer.from_pretrained('bert-base-cased')
#model = BertModel.from_pretrained('bert-base-cased')
#tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
#model = BertModel.from_pretrained('bert-base-uncased')


######
#讀取去除停用詞的問題集
# 读取question的word_total_list.json
import json
with open('clean_q_word_list.json','r') as flie:
    clean_q_word_list = flie.read()
    clean_q_list = json.loads(clean_q_word_list)





####
# 句子的分詞
def get_tokenized_text(sentence):
    
    
    ''''time test
    #計算程序運行時間
    print("      -----------get_tokenized_text---------------    ")
    starttime = time.time()
    print("      starttime:",starttime)
    '''
    # 加入句子开头标签【CLS】和结尾标签【SEP】
    #marked_text = "[CLS] " + sentence + " [SEP]"
    marked_text = "[CLS] "+str(sentence)+" [SEP]"
    #print(marked_text)
    # 把句子拆成多个分词
    tokenized_text_list = tokenizer.tokenize(marked_text)
    
    #print(tokenized_text_list)
    
    
    '''time test
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("      endtime:",endtime)
    print('      time elapsed: ' , timeelapsed , ' seconds\n')
    if workbook:
        global get_tokenized_text_time
        get_tokenized_text_time+=timeelapsed
    '''
    
    return tokenized_text_list
#get_tokenized_text("為甚麼是你?")

####
#整理句子的表示，Bert的問題輸入需要加上ID
# 设定句子的ID
#@jit()
def get_Segment_ID(i,tokenized_text_list):
    segments_ids = [i] * len(tokenized_text_list)
    
    return segments_ids


####
#整理句子的表示，Bert的單詞輸入需要加上ID
#設定單字的ID。使用詞彙表將標記字符串（或標記序列）轉換為單個整數ID（或ID序列）。
def get_indexed_tokens(tokenized_text_list):
    indexed_tokens = tokenizer.convert_tokens_to_ids(tokenized_text_list)
    
    return indexed_tokens



####
#Bert透過PyTorch建立，需要將張量轉成torch張量
# BERT PyTorch要求张量要转成torch张量 Convert inputs to PyTorch tensors
#@jit()
get_tokens_segments_tensor_time=0
def unzip_func(a, b):  
    return a, b 
def get_tokens_segments_tensor(i,sentence):
    
   
    
    #i=args[0]
    #sentence=args[1]
    # 取得句子分词
    tokenized_text = get_tokenized_text(sentence)
    
    #取得indexed_tokens,segments_ids
    indexed_tokens = get_indexed_tokens(tokenized_text)
    segments_ids = get_Segment_ID(i,tokenized_text)
    
    # 转换成pytorch张量
    tokens_tensor = torch.tensor([indexed_tokens])
    segments_tensors = torch.tensor([segments_ids])
    
    
    
    
    return tokens_tensor, segments_tensors




#為了加快運算速度關閉梯度計算
## Torch.no_Grad关闭梯度计算，节省内存，并加快计算速度
def get_encoded_layers(tokens_tensor,segments_tensors):
    with torch.no_grad():
        encoded_layers, _ = model(tokens_tensor, segments_tensors)
    return(encoded_layers)

    

######
#取得Bert 的詞向量
#@jit()
get_sentence_embedding_time=0
def get_sentence_embedding(tokenized_text,encoded_layers):
    
   
    
    
    # Convert the hidden state embeddings into single token vectors
    # Holds the list of 12 layer embeddings for each token
    # Will have the shape: [# tokens, # layers, # features]
    token_embeddings = [] 
    # For each token in the sentence...
    for token_i in range(len(tokenized_text)):
  
      # Holds 12 layers of hidden states for each token 
        hidden_layers = [] 
  
      # For each of the 12 layers...
        for layer_i in range(len(encoded_layers)):
            batch_i = 0
    
        # Lookup the vector for `token_i` in `layer_i`
            vec = encoded_layers[layer_i][batch_i][token_i]
        
            hidden_layers.append(vec)
    
        token_embeddings.append(hidden_layers)
        # Sanity check the dimensions:
        #print ("Number of tokens in sequence:", len(token_embeddings))
        #print ("Number of layers per token:", len(token_embeddings[0]))
        
    ## 特征值[number_of_tokens, 3072]，透过torch.cat把向量拼接
    #concatenated_last_4_layers = [torch.cat((layer[-1], layer[-2], layer[-3], layer[-4]), 0) for layer in token_embeddings]
    ## 特征值相加[number_of_tokens, 768]
    #summed_last_4_layers = [torch.sum(torch.stack(layer)[-4:], 0) for layer in token_embeddings]
    
    ## 平均每个token的倒数第二层，产生一个768长度句向量
    sentence_embedding = torch.mean(encoded_layers[11], 1)
        
   
    
    return sentence_embedding

######
#通過average pooling單詞向量來實現句子的向量
import numpy as np

# 将句子拆成每一个词，将词向量平均建构句子的向量
X_bert = []
i = 0
initial_numpy = np.zeros((768,))
def get_tokens_segments_tensor_(i,sentence):
    '''time test
    #計算程序運行時間
    print("      -----------get_tokens_segments_tensor---------------    ")
    starttime = time.time()
    print("      starttime:",starttime)
    '''
    # 取得句子分词
    tokenized_text = get_tokenized_text(sentence)
    
    #取得indexed_tokens,segments_ids
    indexed_tokens = get_indexed_tokens(tokenized_text)
    segments_ids = get_Segment_ID(i,tokenized_text)
    
    # 转换成pytorch张量
    tokens_tensor = torch.tensor([indexed_tokens])
    segments_tensors = torch.tensor([segments_ids])
    
    
    '''time test
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("      endtime:",endtime)
    print('      time elapsed: ' , timeelapsed , ' seconds\n')
    
    if workbook:
        global get_tokens_segments_tensor_time
        get_tokens_segments_tensor_time+=timeelapsed
    '''
    
    
    return tokens_tensor, segments_tensors
for sentence in clean_q_list:
    # 句子的分词
    tokenized_text = get_tokenized_text(sentence)
    # BERT PyTorch要求张量要转成torch张量 Convert inputs to PyTorch tensors
    tokens_tensor, segments_tensors = get_tokens_segments_tensor_(i,sentence)
    # 取得隐藏层层数
    encoded_layers = get_encoded_layers(tokens_tensor,segments_tensors)

    # 取得单词向量，取平均成句子的向量
    sentence_embedding = get_sentence_embedding(tokenized_text,encoded_layers)
    
    # 把torch张量转成array
    sentence_embedding_array = np.array(sentence_embedding)

    #array全部组合起来
    initial_numpy = np.vstack((initial_numpy,sentence_embedding_array))

bert_vectorizer = initial_numpy

#######
# bert的vectorizer存成h5py
import h5py
# Create a new file
f = h5py.File('bert_vectorizer.h5', 'w')
f.create_dataset('bert_vectorizer', data=bert_vectorizer)
f.close()


## 读取bert的vectorizer
#f = h5py.File('bert_vectorizer.h5', 'r')
#X = f['bert_vectorizer']
#n = np.array(X)
#f.close()

###############
#建立倒排表
# 读取question的word_total_list.json
import json
with open('clean_q_word_list.json','r') as flie:
    clean_q_word_list = flie.read()
    clean_q_list = json.loads(clean_q_word_list)

# 读取question的不重复单字word_total_list.json
with open('q_word_list.json','r') as flie:
    q_word_total_list = flie.read()
    word_total_list = json.loads(q_word_total_list)
# 大写变成小写
word_total = [item.lower() for item in word_total_list]



invert_index=dict()
k = 0  # 提醒我目前已经建立到第几个数据
#print(word_total)
for b in word_total:
    temp=[]
    
    j = 0
    while j < len(clean_q_list):
        

        field=clean_q_list[j]
        
        split_field=field.split()
        #print(split_field)
        
        if b in split_field:
            temp.append(j)
        #print(temp)
        j += 1
       
    k += 1
    if k % 10000 == 0:
        print("k:",k)
    
    invert_index[b]=temp
    #print(invert[b])


# 全部单词(53157个单词)对应问题的字典存成json
import json
file_name_6 = 'inverted_idx.json'
with open(file_name_6,'w') as file_object:
    json.dump(invert_index,file_object)
    
 
'''
#######
#語義相似度
# 加载转化后的文件
from gensim.test.utils import get_tmpfile
from gensim.models import KeyedVectors
from gensim.models.word2vec import Word2Vec

word2vec_file = get_tmpfile("C:\\Users\\liyan\Project 1\\word2vec.txt")
model = KeyedVectors.load_word2vec_format(word2vec_file)

# 获取所有词向量表
word_list = []
for word in model.wv.vocab.keys():  #1.model.wv.vocab.keys()=[[], [], []] 2.model.wv.vocab.keys()=[,,,,]
    word_list += [word]    #1.word_list=[[], [], [], []] 2.word_list=[, , , , , ]
#print(len(word_list))

similar_list = []
# 获取每个词的前10大相关
for word_temp in word_list:

    #建构成字典
    similar_word = []

    #取模型最相似的前10个单词和向量
    similar_top =  model.most_similar(word_temp)

    for similar_top_word in similar_top:
        similar_word += [similar_top_word[0]]  #similar_word=[,,,,,]

    similar_list += [similar_word]   #similar_list=[[,,,,],[,,,],[,,,]]

#转换成字典
from collections import defaultdict 

zip_list = zip(word_list , similar_list) #(word_list[0], [,,,10個]),(word_list[1], [,,,10個]), ...
d = defaultdict(list) #defaultdict(list)意思為d={"":[,,,,10個], "":[,,,,10個], "":[,,,,10個]}
for key, value in zip_list:
    d[key].append(value)   #{key1:[,,,,,10個], key2:[,,,,,10個], ket3=[,,,,,10個]}


##########################保存成txt檔####################################

# 相关词字典写成txt
import codecs
file = codecs.open('related_words.txt', 'w','utf-8-sig') 

# 将每项元素的key和value分拆组成字符串，添加分隔符和换行符
for k,v in d.items():
	file.write(str(k)+' '+str(v)+'\n')
	
# 注意关闭文件
file.close()

##########################保存成json檔####################################
# 相关词字典存成json
import json 

file_name = 'related_words.json'
with open(file_name,'w') as file_object:
    json.dump(d , file_object)
'''


#Bert + 倒排表搜索
#(1)去除停用词
def q_movestopwords(sentence):
    #time test
    print("----------------q_movestopwords-----------------")
    starttime = time.time()
    print("starttime:",starttime)
    
    
    #from nltk.corpus import stopwords
    #from nltk.tokenize import word_tokenize
    import jieba
    
    # 英文的大写全部改成小写
    #q_list_lower = sentence.lower()    

    # 把问题拆解成一个一个单词
    word_tokens = jieba.cut_for_search(sentence)
    
    # 过滤掉停用词
    ## 增加stopwords
    #stop_words = set(stopwords.words('chinese'))
    #stop_words.update(['.', ',', '"', "'", '?', '!', ':', ';', '(', ')', '[', ']', '{', '}'])
    #filtered_words = [word for word in word_tokens if word not in stop_words]
    filtered_words =[word for word in word_tokens if word not in stopwords_list]
    #print(filtered_words)

    # 将单词组合成句子
    filtered_sentens = ' '.join(filtered_words)
    q_filtered_list = [filtered_sentens]
    
    
    '''
    # 过滤掉停用词
    ## 增加stopwords
    stop_words = set(stopwords.words('english'))
    stop_words.update(['.', ',', '"', "'", '?', '!', ':', ';', '(', ')', '[', ']', '{', '}'])
    filtered_words = [word for word in word_tokens if word not in stop_words]

    # 将单词组合成句子
    filtered_sentens = ' '.join(filtered_words)       
    q_filtered_list = [filtered_sentens]
    '''
    
    
    #time test
    endtime = time.time()
    timeelapsed=round(endtime-starttime, 4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    
    
    return q_filtered_list , filtered_words

'''
####
#(2)取得语义相关词
def get_related_words(sentence):
    import json
    # 读取word的相似词字典.json
    with open('related_words.json', 'r') as fp:
        related_words_dic = json.load(fp)
        
    simi_list = []
    for w in sentence:
        temp = related_words_dic[w]
        temp_list = temp[0]  #取前10大相關中的第一個
        simi_list += temp_list

    return simi_list
'''

####
#(3)透过倒排表搜索
def get_inverted_idx(q_total_word):
    
    #time test
    print("----------------get_inverted_idx-----------------")
    starttime = time.time()
    print("starttime:",starttime)
  
    
    
    
    import json
    # 读取word的倒排表字典.json
    with open('inverted_idx.json', 'r') as fp:
        inverted_idx_dic = json.load(fp)
  
    
    # 查找问题的index列表
    q_total_list = []
    for q_word in q_total_word:
    
        # 只找字典中有的单词进行查找
        if q_word in inverted_idx_dic:
            q_total_list += inverted_idx_dic[q_word]

    # 清除重复的问题的index        
    q_total_clean_list = list(set(q_total_list))####


    #time test
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')


    global col
    col+=1
    if workbook:
        table.cell(row=row_,column=col,value=str(timeelapsed))
 


    return q_total_clean_list


####
#(4) 透过index找出所有相对应的问题
def get_total_q(index_q):
    
    #time test
    print("----------------get_total_q-----------------")
    starttime = time.time()
    print("starttime:",starttime)
    
    
    
    # 读取question的word_total_list.json
    import json
    with open('clean_q_word_list.json','r') as flie:
        clean_q_word_list = flie.read()
        clean_q_list = json.loads(clean_q_word_list)
        
    #找到问题与问题集相似的所有问题
    total_q_list = []
    for q in index_q:
        total_q_list += [clean_q_list[q]]
    
    
    
    #time test
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    
    
    
    return total_q_list


######
#(5)执行步骤整理
#get_preprocess:藉由一個使用者問的問題，找到(衍生)58個可能相似的問題 
def get_preprocess(sentence):
    
    #time test
    #計算程序運行時間
    starttime = time.time()
  
    
    
    #1.去除停用詞 2.增加相似詞彙 3.倒排表查詢可能問題
    # 整理完的问题集
    clean_q , word_list = q_movestopwords(sentence)
    #print(clean_q)
    #print(word_list)
    # 相似词列表
    simi_list=word_list
    #simi_list = get_related_words(word_list)######
    # 倒排表找出所有可能的问题index
    q_simi_list = get_inverted_idx(simi_list)
    #print("q_simi_list:",q_simi_list)
    
    # 透过index找出所有相对应的问题
    q_total_list = get_total_q(q_simi_list)
    #print("q_total_list:",q_total_list)
    
    
    
    #time test
    print("----------------get_preprocess-----------------")
    print("starttime:",starttime)
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
   

    return q_total_list , q_simi_list


#(6)Bert轉換成矩陣array
#單詞、句子整理成Bert輸入格式
# 句子的分词
'''268重複
def get_tokenized_text(sentence):
    # 加入句子开头标签【CLS】和结尾标签【SEP】
    marked_text = "[CLS] " + sentence + " [SEP]"
    # 把句子拆成多个分词
    tokenized_text_list = tokenizer.tokenize(marked_text)
    
    return tokenized_text_list
'''    


    
'''重複307
# BERT PyTorch要求张量要转成torch张量 Convert inputs to PyTorch tensors
@jit()
def get_tokens_segments_tensor(i,sentence):
    # 取得句子分词
    tokenized_text = get_tokenized_text(sentence)
    
    #取得indexed_tokens,segments_ids
    indexed_tokens = get_indexed_tokens(tokenized_text)
    segments_ids = get_Segment_ID(i,tokenized_text)
    
    # 转换成pytorch张量
    tokens_tensor = torch.tensor([indexed_tokens])
    segments_tensors = torch.tensor([segments_ids])
    
    return tokens_tensor, segments_tensors
'''



'''重複296
# 设定单字的ID
@jit()
def get_indexed_tokens(tokenized_text_list):
    indexed_tokens = tokenizer.convert_tokens_to_ids(tokenized_text_list)
    
    return indexed_tokens
'''    


'''重複287
# 设定句子的ID
def get_Segment_ID(i,tokenized_text_list):
    segments_ids = [i] * len(tokenized_text_list)
    
    return segments_ids
'''


## Torch.no_Grad关闭梯度计算，节省内存，并加快计算速度
def get_encoded_layers(tokens_tensor,segments_tensors):
    
    
   
    
    
    import torch
    #from pytorch_pretrained_bert import BertTokenizer
    from pytorch_pretrained_bert import BertModel
    #from pytorch_pretrained_bert import BertForMaskedLM
    # 使用bert-base-chinese
    #tokenizer = BertTokenizer.from_pretrained('.\model')#.\model
    model = BertModel.from_pretrained('.\model')
    #tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
    #model = BertModel.from_pretrained('bert-base-uncased')

    with torch.no_grad():
        encoded_layers, _ = model(tokens_tensor, segments_tensors)
        
    
    return(encoded_layers)


'''重複336
#透過句子中的單詞向量平均後變成句子向量
# 取得单词向量，取平均成句子的向量
def get_sentence_embedding(tokenized_text,encoded_layers):
    # Convert the hidden state embeddings into single token vectors
    # Holds the list of 12 layer embeddings for each token
    # Will have the shape: [# tokens, # layers, # features]
    token_embeddings = [] 
    # For each token in the sentence...
    for token_i in range(len(tokenized_text)):
  
      # Holds 12 layers of hidden states for each token 
        hidden_layers = [] 
  
      # For each of the 12 layers...
        for layer_i in range(len(encoded_layers)):
            batch_i = 0
    
        # Lookup the vector for `token_i` in `layer_i`
            vec = encoded_layers[layer_i][batch_i][token_i]
        
            hidden_layers.append(vec)
    
        token_embeddings.append(hidden_layers)
        # Sanity check the dimensions:
        #print ("Number of tokens in sequence:", len(token_embeddings))
        #print ("Number of layers per token:", len(token_embeddings[0]))
        
    ## 特征值[number_of_tokens, 3072]，透过torch.cat把向量拼接
    #concatenated_last_4_layers = [torch.cat((layer[-1], layer[-2], layer[-3], layer[-4]), 0) for layer in token_embeddings]
    ## 特征值相加[number_of_tokens, 768]
    #summed_last_4_layers = [torch.sum(torch.stack(layer)[-4:], 0) for layer in token_embeddings]
    
    ## 平均每个token的倒数第二层，产生一个768长度句向量
    sentence_embedding = torch.mean(encoded_layers[11], 1)
        
    return sentence_embedding
'''
initial_numpy = np.zeros((768,)) 

#取得Bert 的句子向量并整理成array(模型之架構: 輸入層->隱藏層1->隱藏層2->.....->輸出層)
import numpy as np
import threading                   ##1
class MyThread(threading.Thread):   ##1
    def __init__(self,i,new_clean_q):
        threading.Thread.__init__(self)
        self.i, self.new_clean_q=i,new_clean_q
        
    def run(self):
        for sentence in self.new_clean_q:
            
            #print("sentence:",sentence)
            # 句子的分词
            tokenized_text = get_tokenized_text(sentence)
            
            
            
            # BERT PyTorch要求张量要转成torch张量 Convert inputs to PyTorch tensors
            tokens_tensor, segments_tensors = get_tokens_segments_tensor(i,sentence)
    
            # 取得隐藏层层数
            encoded_layers = get_encoded_layers(tokens_tensor,segments_tensors)
    
            # 取得单词向量，取平均成句子的向量
            sentence_embedding = get_sentence_embedding(tokenized_text,encoded_layers)
        
            # 把torch张量转成array
            sentence_embedding_array = np.array(sentence_embedding[0])
            
            #array全部组合起来
            #lock.acquire()
            global initial_numpy         
            initial_numpy= np.vstack((initial_numpy,sentence_embedding_array))
            #print("initial_numpy:",initial_numpy)
            #lock.release()
#lock=threading.Lock()


def get_new_bert(new_clean_q):
    #計算程序運行時間
    print("----------------get_new_bert-----------------")
    starttime = time.time()
    print("starttime:",starttime)
    
    #import numpy as np
    # 将句子拆成每一个词，将词向量平均建构句子的向量
    #X_bert = []
    import numpy as np
    #i = 0
    #initial_numpy = np.zeros((768,))
    #new_clean=new_clean_q
    new_clean_q=np.array(new_clean_q)
    
    

    
    t5=[]
    seperated_array=np.array_split(new_clean_q,3)
    for i in range(3):

        tn=MyThread(i,seperated_array[i])
        #print("seperated_array[i]:",seperated_array[i])
        t5.append(tn)
    #print("t5:",t5)
    
    for j in t5:
        j.start()
        j.join()
    
    #t5[0].start()
    #print("initial_numpy:",initial_numpy)
    bert_vectorizer = initial_numpy
    #print("bert_vectorizer:",bert_vectorizer)
    #print("len(bert_vectorizer):",len(bert_vectorizer))
    
    global col
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("\nendtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    
    return bert_vectorizer


    
#####
#get_cos_list:將60個問題轉變成變成數值(此數值經由餘弦函數得出)
#(7)整理成餘弦函數輸入的格式 (這裡我一共整理成3種文本表示,前兩篇文章為介紹TF-IDF、Word2Vec有興趣的小伙伴可以找一下)
# 输入tfidf、word2vec、bert资料、1次batch的问题量、是哪一种函数(tfidf、word2vec、bert)
def get_cos_list(new_q_list, q_list_batch, model_type):

    #計算程序運行時間
    print("----------------get_cos_list-----------------")
    starttime = time.time()
    print("starttime:",starttime)

    
    import numpy as np
    
    #设定1000笔作batch
    num = 0
    add_num = q_list_batch
    
    #设定空array
    cos_array = np.empty(shape=[0, 1])
    
    # 输入的资料依照是哪一种函数整理成array(分别为询问的问题、全部的问题、共有多少题)
    new_q_array , ask_q_array , array_length = get_arry(new_q_list, model_type)
    #print("new_q_array:",new_q_array)

    while num< array_length:
        if num + add_num  < array_length:
            cos_distance_array = get_cos_distance(ask_q_array, new_q_array[num : num + add_num])
            cos_array = np.append(cos_array , cos_distance_array)
        else:
            cos_distance_array = get_cos_distance(ask_q_array, new_q_array[num : ])
            cos_array = np.append(cos_array , cos_distance_array)
            
        num += add_num
    #print("cos_distant:",cos_array)
    
    

    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    return cos_array   


####
#(8)Bert整理成array
def get_arry(data, model_type):

    if model_type == 'bert':
        import numpy as np
        
        new_array = np.array(data)
        ask_q_array = np.array(data[-1:])
        length = len(data)

    else:
        print('no input model')
    
    return new_array,ask_q_array,length

#(9)余弦函数计算
def get_cos_distance(X1, X2):
    
    import tensorflow as tf
    #import numpy as np    
    
    #sess=tf.Session()
    with tf.Session(config=tf.ConfigProto(allow_soft_placement=True, log_device_placement=True)) as sess:
        with tf.device('/cpu:0'):
            # calculate cos distance between two sets
            # more similar more big
    
            # 新增询问的问题
            X1_matrix = X1
            
            # 全部的问题
            X2_matrix = X2
          
            
            # 计算矩阵型态
            (k,n) = X1_matrix.shape
            (m,n) = X2_matrix.shape
        with tf.device('/gpu:0'):
            
    
            # 求模
            X1_norm = sess.run(tf.sqrt(tf.reduce_sum(tf.square(X1_matrix), axis=1)))
            X2_norm = sess.run(tf.sqrt(tf.reduce_sum(tf.square(X2_matrix), axis=1)))
            # 内积
            X1_X2_muti = sess.run(tf.matmul(X1_matrix, tf.transpose(X2_matrix)))
            X1_X2_norm = sess.run(tf.matmul(tf.reshape(X1_norm,[k,1]),tf.reshape(X2_norm,[1,m])))
            # 计算余弦距离
            #cos_dis = get_cos_distance(X1_X2_muti, X1_X2_norm )
            
            cos_dis = sess.run(tf.div(X1_X2_muti, X1_X2_norm))
            #cos = cos_dis.eval(session=session)

    sess.close()
    return cos_dis


###
#(10)找出前index_num個最相似的問题，返回前58個相似問題的index。index_num個最相似的問题<->前58個相似問題中對應的index
def get_top_index(cos_array,index_num):
 
    #計算程序運行時間
    print("----------------get_top_index-----------------")
    starttime = time.time()
    print("starttime:",starttime)
    
    
    import heapq
    # array转list
    cos_list = cos_array.tolist()
    
    # 最大的索引个数(因为list最后就是新增的问题要去掉,所以要多取1个)
    #因為尾巴新增的問題(使用者問得問題)一定是數值最高的(100%=1符合)，故最大索引個數要加1
    top_num_count = index_num+1
    #map(cos_list.index, heapq.nlargest(top_num_count, cos_list)):
    #利用cos_list.index(heapq.nlargest(top_num_count, cos_list))分別找出
    #前index_num對應數值之index
    max_index =map(cos_list.index, heapq.nlargest(top_num_count, cos_list))#max_index=[相對應index, 相對應index]
    #print(max_index)
    #print("heapq.nlargest(top_num_count, cos_list):",heapq.nlargest(top_num_count, cos_list))
    #heapq.nlargest(n, iterable, key=None):
    #从 iterable 所定义的数据集中返回前 n 个最大元素组成的列表。 如
    #果提供了 key 则其应指定一个单参数的函数，用于从 iterable 的每个
    #元素中提取比较键 (例如 key=str.lower)。 等价于: 
    #sorted(iterable, key=key, reverse=True)[:n]。
    
    #产生成list
    top_index_list = list(max_index)
    top_index = top_index_list
    #print("top_idxs:",top_index)
    
    
    endtime = time.time()
    timeelapsed=round(endtime-starttime,4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    
    return top_index

######
#(11)找出相对应最有可能的答案
def get_answer(a_index , q_index):

    
    #計算程序運行時間
    print("----------------get_answer-----------------")
    starttime = time.time()
    print("starttime:",starttime)
    
    
    
    import json
    #print("time now:",time.time())
    with open('a_list.json', 'r') as file:
        answers = file.read()
        alist = json.loads(answers)
    
    q_index_list = []
    #print(q_index_list)
    #从新建立好的倒排表list反推可能问题list
    for u in a_index:   #q_index不包括新增的問題(使用者問得問題)，len=58
        if u < len(q_index):  #如果前index_num個相似問題之index在前58個相似問題之
                              #長度內，當超過範圍時(第59個:使用者問的問題)，即不加入
            q_index_list += [q_index[u]]
    #print("q_index_list:",q_index_list)
    

    #从可能问题的list找到相对应的答案list
    top_a = []
    
    for i in q_index_list:
        top_a += [alist[i]]
    #print("alist:",top_a)
    
    
    endtime = time.time()
    timeelapsed=round(endtime-starttime, 4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    
    
    return top_a


####
#(12)执行Bert的步骤整理表
#@jit()
def get_top_results_bert(query):
    #計算程序運行時間
    starttime = time.time()
    
    
    
    # top_idxs存放相似度最高的（存在qlist里的）问题的下表 
    # hint: 利用priority queue来找出top results. 思考为什么可以这么做？ 
    
    """
    给定用户输入的问题 query, 返回最有可能的TOP 5问题。这里面需要做到以下几点：
    1. 利用倒排表来筛选 candidate （需要使用related_words). 
    2. 对于候选文档，计算跟输入问题之间的相似度
    3. 找出相似度最高的top5问题的答案
    """
    # 找出与问题与问题集中相似的问题
    preprocess_q, q_index_list = get_preprocess(query) #preprocess_q=[,,,,可能的問題(文字)]   q_index_list=[,,,可能的問題的index]
    # 建立新的bert的句子词向量(单词向量平均)
    q_new_list = preprocess_q + [query]  #加入使用者問的問題，len長度變59
    new_bert = get_new_bert(q_new_list)
    #取得新问题对问题集的Bert余弦函数
    cos_distant = get_cos_list(new_bert, q_list_batch = 1000, model_type = 'bert')
    # 取得问题在问题集中最相近的5个问题
    top_idxs = get_top_index(cos_distant,5)

    #取得最接近的答案list
    alist = get_answer(top_idxs , q_index_list)

    ##显示答案
    #for top_a in alist:
        #print(top_a) 
        
    global row_
    print("----------------get_top_results_bert-----------------")
    print("starttime:",starttime)
    endtime = time.time()
    timeelapsed=round(endtime-starttime, 4)
    print("endtime:",endtime)
    print('time elapsed: ' , timeelapsed , ' seconds\n')
    
    global col
    col+=1
    table.cell(row=row_,column=col,value=str(timeelapsed))
    
    #row_+=1
    #col=1

    return alist  # 返回相似度最高的问题对应的答案，作为TOP5答案





'''
col =1
row_=2
if(workbook==None):  #建立xlsx檔
    write_excel_xlsx("test.xlsx","test_table",
                     [["Question","q_movestopwords","get_inverted_idx","get_total_q",
                      "get_preprocess","get_tokenized_text","get_tokens_segments_tensor",
                      "get_encoded_layers","get_sentence_embedding","sentences","get_new_bert",
                      "get_cos_list","get_top_index","get_answer","總共"]])

workbook = openpyxl.load_workbook("test.xlsx") #載入xlsx檔
sheetnames = workbook.get_sheet_names()
table=workbook.get_sheet_by_name(sheetnames[0])
table=workbook.active
####
#(13)测试使用Bert的结果
#问题测试   
#loop start
test_query1 = "南明的首都在哪？"
table.cell(row=row_,column=col,value=test_query1)
#test_query2 = "What are doctors part of?"
answer=get_top_results_bert(test_query1)
print ('bert:',answer)
print('answer:',answer[0])
workbook.save('test.xlsx')
#loop end
'''




#測試1
#讀取json
import json
problems=[]
with open('.\\data\\DRCD_\\DRCD_training.json', 'r',encoding="utf-8") as file:
    jsonfile = json.loads(file.read())
paragraphs=jsonfile["data"][0]["paragraphs"]
for paragraph_ in paragraphs:
    #print(paragraph_)
    qas=paragraph_["qas"]
    #print(qas)
    for qas_ in qas:
        #print(qas_["question"])
        problems.append(qas_["question"])
#print(len(problems))
c=0
col =1
row_=2
if(workbook==None):  #建立xlsx檔
    write_excel_xlsx("testV4.xlsx","test_table",
                             [["Question","q_movestopwords","get_inverted_idx","get_total_q",
                              "get_preprocess","get_new_bert","get_cos_list","get_top_index",
                              "get_answer","總共","正確答案"]])
workbook = openpyxl.load_workbook("testV4.xlsx") #載入xlsx檔
sheetnames = workbook.get_sheet_names()
table=workbook.get_sheet_by_name(sheetnames[0])
table=workbook.active
for test_query1 in problems:
    c+=1
    #loop start
    table.cell(row=row_,column=col,value=test_query1)
    #test_query2 = "What are doctors part of?"
    answer=get_top_results_bert(test_query1)
    print ('bert:',answer)
    print('answer:',answer[0])
    #global col
    #global row_
    col+=1
    table.cell(row=row_,column=col,value=str(answer[0]))
    row_+=1
    col=1
    workbook.save('testV4.xlsx')
    #loop end




'''
#print ('bert:',get_top_results_bert(test_query2))
#print("answer:",get_top_results_bert("哈囉"))
answer=get_top_results_bert("哈囉")
if len(answer):
    cmd =answer[0]
    print(cmd)
else:
    cmd="很抱歉無法回答你的問題"
    print(cmd)
'''


'''
#(1)驗證
#get_preprocess("廣州的快速公交運輸系統每多久就會有一輛巴士?") 
q_index_list=get_preprocess("廣州的快速公交運輸系統每多久就會有一輛巴士?")[1]
q_new_list = get_preprocess("廣州的快速公交運輸系統每多久就會有一輛巴士?")[0] + ["廣州的快速公交運輸系統每多久就會有一輛巴士?"]
new_bert = get_new_bert(q_new_list)
#print(len(get_preprocess("廣州的快速公交運輸系統每多久就會有一輛巴士?")[1]))
#print("len(new_bert):",new_bert)
#print(len(new_bert))
cos_distant = get_cos_list(new_bert, q_list_batch = 1000, model_type = 'bert')
#print("len(cos_distant):",len(cos_distant))
top_idxs = get_top_index(cos_distant,5)
#print(top_idxs)
alist = get_answer(top_idxs , q_index_list)  #出現四個可能答案，少一個不知道
                                             #為甚麼，但第一個確定是最接近的答案
print(alist)
'''


'''
#(2)驗證
get_preprocess("從哪一天開始在廣州市內騎摩托車會被沒收？")
q_index_list=get_preprocess("從哪一天開始在廣州市內騎摩托車會被沒收？")[1]
q_new_list = get_preprocess("從哪一天開始在廣州市內騎摩托車會被沒收？")[0] + ["從哪一天開始在廣州市內騎摩托車會被沒收？"]
new_bert = get_new_bert(q_new_list)
#print(len(get_preprocess("從哪一天開始在廣州市內騎摩托車會被沒收？")[1]))
#print("len(new_bert):",new_bert)
#print(len(new_bert))
cos_distant = get_cos_list(new_bert, q_list_batch = 1000, model_type = 'bert')
#print("len(cos_distant):",len(cos_distant))
top_idxs = get_top_index(cos_distant,5)
#print(top_idxs)
alist = get_answer(top_idxs , q_index_list)  #出現四個可能答案，少一個不知道
                                             #為甚麼，但第一個確定是最接近的答案
print(alist)
'''

'''
#(3)驗證方式邏輯
get_preprocess("廣州白雲國際機場在完成第三條跑道的後八年哪一座機場也會有第三跑道？")
q_index_list=get_preprocess("廣州白雲國際機場在完成第三條跑道的後八年哪一座機場也會有第三跑道？?")[1]
q_new_list = get_preprocess("廣州白雲國際機場在完成第三條跑道的後八年哪一座機場也會有第三跑道？")[0] + ["廣州白雲國際機場在完成第三條跑道的後八年哪一座機場也會有第三跑道？"]
new_bert = get_new_bert(q_new_list)
print(len(get_preprocess("廣州白雲國際機場在完成第三條跑道的後八年哪一座機場也會有第三跑道？")[1]))
print("len(new_bert):",new_bert)
print(len(new_bert))
cos_distant = get_cos_list(new_bert, q_list_batch = 1000, model_type = 'bert')
print("len(cos_distant):",len(cos_distant))
top_idxs = get_top_index(cos_distant,5)
print(top_idxs)
alist = get_answer(top_idxs , q_index_list)  #出現四個可能答案，少一個不知道
                                             #為甚麼，但第一個確定是最接近的答案
print(alist)
'''





'''
import json
problems=[]
with open('.\\data\\DRCD_\\DRCD_training.json', 'r',encoding="utf-8") as file:
    jsonfile = json.loads(file.read())
paragraphs=jsonfile["data"][0]["paragraphs"]
for paragraph_ in paragraphs:
    #print(paragraph_)
    qas=paragraph_["qas"]
    #print(qas)
    for qas_ in qas:
        #print(qas_["question"])
        problems.append(qas_["question"])
#print(len(problems))

c=0
col =1
row_=2
if(workbook==None):  #建立xlsx檔
    write_excel_xlsx("test.xlsx","test_table",
                             [["Question","q_movestopwords","get_inverted_idx","get_total_q",
                              "get_preprocess","get_tokenized_text","get_tokens_segments_tensor",
                              "get_encoded_layers","get_sentence_embedding","sentences","get_new_bert",
                              "get_cos_list","get_top_index","get_answer","總共"]])
workbook = openpyxl.load_workbook("test.xlsx") #載入xlsx檔
sheetnames = workbook.get_sheet_names()
table=workbook.get_sheet_by_name(sheetnames[0])
table=workbook.active
############
#和android通訊
#獲取IP地址方法一
import socket
def get_host_ip():
    """
    查詢本機ip地址
    :return: ip
    """
    try:
        s=socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip=s.getsockname()[0]
    finally:
        s.close()
    
    return ip
print(socket.gethostname())
print(get_host_ip())
#測試是否可連接成功
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
hostip=get_host_ip()
port = 9999
s.bind((hostip, port))
s.listen(4)
while True:
    conn, addr = s.accept()
    print('Connected by', addr)
def socket_server_for_poc():
    hostip = get_host_ip()
    print('host ip', hostip)  # 应该显示为：127.0.1.1
    port = 9999
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind((hostip, port))
    s.listen(4)
    while True:
        
        conn, addr = s.accept()
        print('Connected by', addr)
        data = conn.recv(1024)
        #print('Received', repr(data.decode("utf-8")))
        print('Question:', repr(data.decode("utf-8")))
        if not data:
            break
        
        question=repr(data.decode("utf-8"))
        answer=get_top_results_bert(question)
        print ('Answer:',answer)
        
        #cmd = input("Please intput your cmd:")
        if len(answer):
            cmd =answer[0]
        else:
            cmd="很抱歉無法回答你的問題"
            
        conn.sendall(cmd.encode())  # 发送
        print('send', cmd)
        conn.close()
socket_server_for_poc()
'''


'''
import json
problems=[]
with open('.\\data\\DRCD_\\DRCD_training.json', 'r',encoding="utf-8") as file:
    jsonfile = json.loads(file.read())
paragraphs=jsonfile["data"][0]["paragraphs"]
for paragraph_ in paragraphs:
    #print(paragraph_)
    qas=paragraph_["qas"]
    #print(qas)
    for qas_ in qas:
        #print(qas_["question"])
        problems.append(qas_["question"])
#print(len(problems))

c=0
col =1
row_=2
if(workbook==None):  #建立xlsx檔
    write_excel_xlsx("test.xlsx","test_table",
                             [["Question","q_movestopwords","get_inverted_idx","get_total_q",
                              "get_preprocess","get_tokenized_text","get_tokens_segments_tensor",
                              "get_encoded_layers","get_sentence_embedding","sentences","get_new_bert",
                              "get_cos_list","get_top_index","get_answer","總共"]])
workbook = openpyxl.load_workbook("test.xlsx") #載入xlsx檔
sheetnames = workbook.get_sheet_names()
table=workbook.get_sheet_by_name(sheetnames[0])
table=workbook.active
#獲取IP地址方法二
import socket
def getipaddrs(hostname):#只是为了显示IP，仅仅测试一下
    result = socket.getaddrinfo(hostname, None, 0, socket.SOCK_STREAM)
    return result[2][4][0]
print(socket.gethostname())
print(getipaddrs(socket.gethostname()))
#測試是否可連接成功
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
hostip = getipaddrs(hostname)
port = 9999
s.bind((hostip, port))
s.listen(4)
while True:
    conn, addr = s.accept()
    print('Connected by', addr)
def socket_server_for_poc():
    host = '127.0.0.1'  # 为空代表为本地host
    hostname = socket.gethostname()
    hostip = getipaddrs(hostname)
    print('host ip', hostip)  # 应该显示为：127.0.1.1
    port = 9999
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind((hostip, port))
    s.listen(4)
    while True:
        
        conn, addr = s.accept()
        print('Connected by', addr)
        data = conn.recv(1024)
        #print('Received', repr(data.decode("utf-8")))
        print('Question:', repr(data.decode("utf-8")))
        if not data:
            break
        
        question=repr(data.decode("utf-8"))
        answer=get_top_results_bert(question)
        print ('Answer:',answer)
        
        #cmd = input("Please intput your cmd:")
        if len(answer):
            cmd =answer[0]
        else:
            cmd="很抱歉無法回答你的問題"
            
        conn.sendall(cmd.encode())  # 发送
        print('send', cmd)
        conn.close()

socket_server_for_poc()
'''


